import openpyxl
from openpyxl.workbook.defined_name import DefinedName
import random
import string
import os

# Создание новой книги
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# Заполнение некоторых ячеек для примера
ws['A1'] = 10
ws['A2'] = 20
ws['A3'] = 30

# Генерация случайного имени файла
random_filename = 'workbook_' + ''.join(random.choices(string.ascii_letters + string.digits, k=8)) + '.xlsm'

# Сохранение книги как макрос-enabled (xlsm)
wb.save(random_filename)

# Открытие книги заново для добавления VBA
wb = openpyxl.load_workbook(random_filename, keep_vba=True)
ws = wb.active

# Чтение содержимого VBA модуля
with open('/workspace/vba_module.bas', 'r') as f:
    vba_code = f.read()

# Добавление VBA проекта и модуля
# Примечание: openpyxl имеет ограниченную поддержку VBA. Для более сложных сценариев может потребоваться win32com или xlwings.
# Этот пример показывает базовый подход, но для полноценного переноса модуля может потребоваться ручное редактирование .bas файла или другая библиотека.
# Мы попробуем добавить код напрямую, но это может не сработать полностью корректно без правильной структуры VBA проекта.

# Создание нового модуля VBA в книге
vba_module = wb.create_sheet('Module1')
# openpyxl не позволяет напрямую добавлять код VBA таким образом.
# Необходимо использовать более низкоуровневый доступ или другую библиотеку.
# Попробуем обойти это ограничение, сохранив код в листе, а затем вручную скопируем его в VBA редакторе, если это будет возможно.
# Однако для автоматизации мы должны использовать библиотеку, которая поддерживает VBA, например, pywin32 под Windows.

# Так как полноценная автоматизация сложна с openpyxl, я просто сохраню книгу с данными.
# Для демонстрации функции VBA пользователь должен будет вручную импортировать модуль.

# Сохранение книги
wb.save(random_filename)

print(f"Файл Excel '{random_filename}' создан и сохранен.")
print(f"В него добавлены данные в ячейки A1:A3 (10, 20, 30).")
print("Функция VBA SUM_CELLS была записана в /workspace/vba_module.bas.")
print("Вам нужно вручную импортировать этот модуль в VBA редакторе Excel.")